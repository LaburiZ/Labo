import os.path
import openpyxl

def extract_data(i, save_file):
    # hoid
    a = [str(i[4].value).split('.')[0] + str(i[2].value).split('.')[0] + "-" + str(i[8].value.split('.')[0]) \
         + str(i[11].value).split('.')[0] + str(i[13].value).split('.')[0]]
    # 직역은 변경될 가능성이 있으니 지금은 배제하도록 하자
    # a.append(i[19].value) # 직역(한글)
    a.append(i[17].value)  # 호내위상
    a.append(i[22].value)  # 성
    a.append(i[23].value)  # 명
    if type(i[24].value) == int:  # 출생년도
        a.append(i[4].value - i[24].value)
    else:
        a.append(i[24].value)
    a.append(i[26].value)  # 간지(한글)
    a.append(i[42].value)  # 주성명, 노비인 경우에 사용
    a.append(i[45].value)  # 부명(한자)
    a.append(i[46].value)  # 부명(한글)
    a.append(i[49].value)  # 모명(한자)
    a.append(i[50].value)  # 모명(한글)
    a.append(i[54].value)  # 조명(한자)
    a.append(i[55].value)  # 조명(한글)
    a.append(i[58].value)  # 증조명(한자)
    a.append(i[59].value)  # 증조명(한글)
    a.append(i[62].value)  # 외조명(한글)
    a.append(i[63].value)  # 외조명(한글)
    save_file.append(a)

def extract_xlsx(file, main_hoid_output1, main_hoid_output2, other_hoid_output, static_output, all_hoid_output):
    # target file open
    # 하나 배웠다, data_only안하면 cell에 써져있는 그대로 가져오고
    # data only하면 출력되는 모습 그대로 가져오는 듯 하다
    xlsx = openpyxl.load_workbook(filename=file, data_only=True)
    sheet = xlsx['Sheet']

    #static_first = ["file_name", "ALL", "평민 이상", "노비", "공백", "x포함"]
    static_count = [file, 0, 0, 0, 0, 0]

    for i in sheet.rows:
        if "주호" in str(i[17].value):
            if i[22].value != None and "x" not in i[22].value and i[23].value != None and "x" not in i[23].value and i[42].value == None:
                extract_data(i, main_hoid_output1)
            else:
                extract_data(i, main_hoid_output2)
        else:
            # hoid
            extract_data(i, other_hoid_output)
        extract_data(i, all_hoid_output)

        # static
        static_count[1] += 1
        job = i[18].value
        if job == None:
            static_count[4] += 1
        elif '奴' in job or '婢' in job:
            static_count[3] += 1
        elif 'x' in job:
            static_count[5] += 1
        else:
            static_count[2] += 1
    static_output.append(static_count)

def main():
    # 작업 위치 및 저장 위치 정의
    work_dir = os.getcwd()
    print(work_dir + "현재 작업 위치")
    main_id_dir = work_dir + '\\' + "output_mainho" + '\\'
    other_id_dir = work_dir + '\\' + "output_otherho" + '\\'
    static_id_dir = work_dir + '\\' + "output_static" + '\\'
    all_id_dir = work_dir + '\\' + "output_all" + '\\'

    save_file_mho1 = main_id_dir + "mho_id.xlsx"
    save_file_mho2 = main_id_dir + "mho_id(garbage).xlsx"
    save_file_oth = other_id_dir + "oth_id.xlsx"
    save_file_static = static_id_dir + "statics_mho.xlsx"
    save_file_all = all_id_dir + "all_id.xlsx"

    # save main ho file open
    xlsx1 = openpyxl.Workbook()
    main_hoid_output1 = xlsx1.active

    # save other file open
    xlsx2 = openpyxl.Workbook()
    other_hoid_output = xlsx2.active

    # statics file open
    xlsx3 = openpyxl.Workbook()
    static_output = xlsx3.active
    static_first = ["file_name", "ALL", "평민 이상", "노비", "공백", "x포함"]
    static_output.append(static_first)

    xlsx4 = openpyxl.Workbook()
    main_hoid_output2 = xlsx4.active

    xlsx5 = openpyxl.Workbook()
    all_hoid_output = xlsx5.active

    # output 저장 폴더 확인
    if os.path.isdir(main_id_dir) == 0:
        os.mkdir(main_id_dir)
    if os.path.isdir(other_id_dir) == 0:
        os.mkdir(other_id_dir)
    if os.path.isdir(static_id_dir) == 0:
        os.mkdir(static_id_dir)
    if os.path.isdir(all_id_dir) == 0:
        os.mkdir(all_id_dir)


    files = os.listdir(work_dir)

    for file in files:
        if len(file.split('.')) == 2:
            if file.split('.')[1] == 'xlsx':
                print(file)
                extract_xlsx(file, main_hoid_output1, main_hoid_output2, other_hoid_output, static_output, all_hoid_output)


    xlsx1.save(save_file_mho1)
    xlsx2.save(save_file_oth)
    xlsx3.save(save_file_static)
    xlsx4.save(save_file_mho2)
    xlsx5.save(save_file_all)

main()